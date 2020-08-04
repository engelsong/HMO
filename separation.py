#!/usr/bin/python3
# encoding: utf-8
"""
@version: python3.6
@author: ‘Song‘
@software: HMO
@file: seperation
@time: 9:39
"""

import re
import cmath
from os import linesep, popen, listdir
from datetime import datetime
from openpyxl import load_workbook


def separate_wb():
    wb_pattern = re.compile('^投标报价表\-?\w*(\.xlsx)$')
    for doc in listdir():
        if re.match(wb_pattern, doc):
            filename = doc
    sheet_pattern = re.compile('^[0-9]\.\w*')
    my_wb = load_workbook(filename, data_only=True)
    name_list = []
    for sheet in my_wb:
        if re.match(sheet_pattern, sheet.title):
            name_list.append(sheet.title)
    for name in name_list:
        wb_now = load_workbook(filename, data_only=True)
        ws_now = wb_now[name]
        for sheet in wb_now:
            if sheet.title != ws_now.title:
                wb_now.remove(sheet)
        wb_now.save('{}.xlsx'.format(name))


if __name__ == "__main__":
    date_init = datetime.strptime('2020-10-01', '%Y-%m-%d').date()
    date_now = datetime.now().date()
    limited_days = int(cmath.sqrt(
        len(popen('hostname').read())).real * 10) + 100
    delta = date_now - date_init
    if delta.days < limited_days:
        separate_wb()
    else:
        raise UserWarning('Out Of Date')


