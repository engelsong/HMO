#!/usr/bin/python3
# encoding: utf-8
"""
@version: python3.6
@author: ‘Song‘
@software: HMO
@file: create_training.py
@time: 16:22
"""


import cmath
from os import linesep, popen
from datetime import datetime
from docx import Document
from openpyxl import Workbook
from openpyxl.styles import Border, Side, Alignment, Font, PatternFill
from openpyxl.workbook.properties import CalcProperties


class Project(object):
    """通过Word文档建立项目对象保存项目信息"""
    def __init__(self, document_name):
        self.name = None
        self.code = None
        self.date = None
        self.destination = None
        self.trans = None
        self.totalsum = 0
        self.is_lowprice = False  # 是否为低价法
        self.is_tech = False  # 是否有技术服务
        self.is_qa = False  # 是否有售后
        self.is_cc = False  # 是否来华培训
        self.techinfo = []  # 存放技术服务信息，格式为[人数，天数，[伙食费，住宿费，公杂费]]
        self.training_days = 0  # 来华培训天数
        self.training_num = 0  # 来华培训人数
        self.qc = []  # 法检物资序号
        self.commodities = {}  # 存放物资信息字典
        document = Document(document_name)
        table1, table2 = document.tables  # 读取两个表格
        project_info = []
        for cell in table1.column_cells(1):
            project_info.append(cell.text)
        table2_length = len(table2.rows)
        for index in range(1, table2_length):  # 从第2行开始读取表格
            temp = []
            row_now = table2.row_cells(index)
            length_row = len(row_now)
            for i in range(1, length_row):  # 将每行信息放入暂存数组
                temp.append(row_now[i].text)
            temp.append(row_now[0].text)  # 把物资编号放在最后一位
            self.commodities[index] = temp
        self.name, self.code, self.date, self.destination, self.trans = project_info[0:5]
        self.totalsum = int(project_info[5])
        if project_info[6] in 'yY':
            self.is_lowprice = True
        if project_info[7] in 'yY':
            self.is_tech = True
            self.techinfo += list(map(int, project_info[8:10]))
            self.techinfo.append(list(map(int, project_info[10].split())))
        if project_info[11] in 'yY':
            self.is_qa = True
        if project_info[12] in 'yY':
            self.is_cc = True
            self.training_days = int(project_info[14])  # 读取来华陪训天数
            self.training_num = int(project_info[13])  # 读取来华培训人数
        if project_info[-1] != '':
            self.qc += list(map(int, project_info[-1].split()))
            self.qc.sort()

    def show_info(self):
        print('项目名称:', self.name)
        print('项目代码:', self.code)
        print('开标日期:', self.date)
        print('目的地:', self.destination)
        print('运输方式:', self.trans)
        print('对外货值：', self.totalsum)
        print('是否为低价法', '是' if self.is_lowprice is True else '否')
        print('是否有技术服务:', '是' if self.is_tech is True else '否')
        print('是否有售后服务:', '是' if self.is_qa is True else '否')
        print('是否有来华培训', '是' if self.is_cc is True else '否')
        if self.is_tech:
            print('技术服务人数:', self.techinfo[0])
            print('技术服务天数:', self.techinfo[1])
            print('伙食费:', self.techinfo[2][0])
            print('住宿费:', self.techinfo[2][1])
            print('公杂费:', self.techinfo[2][2])
        if self.is_cc:
            print('来华培训人数：', self.training_num)
            print('来华培训天数：', self.training_days)
        if len(self.qc) > 0:
            print('法检物资：', self.qc)

    def show_commoditiy(self):
        temp_list = sorted(list(self.commodities.keys()))
        for i in temp_list:
            print(i)
            for j in self.commodities[i]:
                print(j)


class Quotation(object):
    """通过project实例创建报价表"""
    def __init__(self, project):
        self.project = project
        self.wb = Workbook()
        self.wb.calculation = CalcProperties(iterate=True)
        self.ws_input = None
        self.ws_cost = None
        self.ws_examination = None
        self.ws_lawexam = None
        self.ws_techserve = None
        self.ws_training = None
        self.ws_selection = None
        self.ws_itemized_quotation = None
        self.ws_summed_quotationn = None
        self.ws_general = None
        self.col_name = ''


    def create_all(self, filename):
        self.create_training()
        self.wb.save(filename)


    def create_training(self):
        """创建来华培训费报价表"""
        self.ws_training = self.wb.create_sheet('5.来华培训费报价表', 3)
        colum_title = ['序号', '费用名称', '', '费用计算方式', '', '', '人民币（元）', '其中含购汇人民币限额']
        title_width = [6, 14, 7, 14, 7, 12, 14, 12]
        # 设置基本的样式
        real_side = Side(style='thin')
        full_border = Border(left=real_side, right=real_side, top=real_side, bottom=real_side)
        slash_border = Border(diagonal=real_side, diagonalDown=True, left=real_side, right=real_side,
                              top=real_side, bottom=real_side)
        ctr_alignment = Alignment(horizontal='center', vertical='center', wrapText=True)
        right_alignment = Alignment(horizontal='right', vertical='center', wrapText=True)
        left_alignment = Alignment(horizontal='left', vertical='center', wrapText=True)
        bold_font = Font(name='宋体', bold=True, size=12)
        normal_font = Font(name='宋体', size=12)
        title_font = Font(name='黑体',bold=True, size=14)
        yellow_fill = PatternFill(fill_type='solid', start_color='FFFF00', end_color='FFFF00')


        # 初始化表格
        colum_number = len(colum_title)
        row_number = 20
        for i in range(colum_number):
            for j in range(1, row_number):  # 第一列留下给表头
                cell_now = self.ws_training.cell(row=j + 1, column=i + 1)
                if j < 17:  # 给主体cell设置样式
                    cell_now.border = full_border
                    cell_now.font = normal_font
                    if i == 6:  # 数字列右对齐
                        cell_now.alignment = right_alignment
                    else:
                        cell_now.alignment = ctr_alignment
                else:  # 最后两行左对齐
                    cell_now.font = normal_font
                    cell_now.alignment = left_alignment
        for i in range(len(title_width)):  # 修改列宽
            self.ws_training.column_dimensions[
                self.ws_training.cell(row=4, column=i + 1).column_letter].width = title_width[i]
        for i in range(2, row_number):  # 修改行高
            self.ws_training.row_dimensions[
                self.ws_training.cell(row=i, column=1).row].height = 20
        self.ws_training.row_dimensions[20].height = 30

        # 打上斜线
        cell_coor = ['D14', 'E14', 'F14']
        for cell in cell_coor:
            self.ws_training[cell].border = slash_border

        # 创建标题行
        self.ws_training['A1'].font = title_font
        self.ws_training['A1'].alignment = ctr_alignment
        index = 4
        if self.project.is_tech:
            index += 1
        self.ws_training['A1'] = '{}.来华培训费报价表'.format(index)
        self.ws_training.row_dimensions[1].height = 30
        self.ws_training.merge_cells('A1:H1')

        # 填写表头
        index = 0
        for i in self.ws_training['A2':'H2'][0]:
            # print(index, i)
            if colum_title[index] != '':
                i.value = colum_title[index]
                i.font = bold_font
            index += 1
        self.ws_training['D3'].value = '标准'
        self.ws_training['E3'].value = '人数'
        self.ws_training['F3'].value = '天（次）数'
        self.ws_training['D3'].font = bold_font
        self.ws_training['E3'].font = bold_font
        self.ws_training['F3'].font = bold_font

        # 填写数据
        col_a = ['一', '二', 1, 2, 3, 4, 5, 6, '三', '四', 1, 2,  '', '五', '']  # 序号
        col_b = ['培训费', '接待费', '日常伙食费', '住宿费', '宴请费', '零用费', '小礼品费', '人身意外伤害保险',
                 '国际旅费', '管理费', '承办管理费', '管理人员费', '', '合计']  # 费用名称
        for index in range(4, 18):  # 填写前两列
            self.ws_training['A{}'.format(index)] = col_a[index - 4]
            if isinstance(col_a[index - 4], int):
                self.ws_training['A{}'.format(index)].alignment = right_alignment
            self.ws_training['B{}'.format(index)] = col_b[index - 4]
        self.ws_training['C15'].value = '伙食费'
        self.ws_training['C16'].value = '住宿费'
        # 填写E列
        for num in range(6, 13):  # 填写培训人数
            self.ws_training['E{}'.format(num)].number_format = '0'
            self.ws_training['E{}'.format(num)].value = self.project.training_num
        self.ws_training['E4'].number_format = '0'
        self.ws_training['E4'].value = self.project.training_num
        self.ws_training['E15'].number_format = '0'
        self.ws_training['E15'].value = self.project.training_num
        self.ws_training['E16'].number_format = '0'
        self.ws_training['E16'].value = self.project.training_num
        # 填写D列
        for num in [4, 6, 7, 9, 15, 16]:
            self.ws_training['D{}'.format(num)].number_format = '0"元/人*天"'
        for num in range(10,13):
            self.ws_training['D{}'.format(num)].number_format = '0"元/人"'
        self.ws_training['D8'].number_format = '0"元/人*次"'
        self.ws_training['D4'] = 320
        self.ws_training['D6'] = 140
        self.ws_training['D7'] = 300
        self.ws_training['D8'] = 150
        self.ws_training['D9'] = 80
        self.ws_training['D10'] = 200
        self.ws_training['D11'] = 100
        self.ws_training['D12'] = 0
        self.ws_training['D15'] = 140
        self.ws_training['D16'] = 300
        # 填写F列
        for i in [4, 6, 7, 8, 9, 10, 11, 12, 15, 16]:
            self.ws_training['F{}'.format(i)].number_format = '0'
        self.ws_training['F4'] = self.project.training_days
        self.ws_training['F6'] = self.project.training_days
        self.ws_training['F7'] = self.project.training_days
        self.ws_training['F8'] = 1
        self.ws_training['F9'] = self.project.training_days
        self.ws_training['F10'] = 1
        self.ws_training['F11'] = 1
        self.ws_training['F12'] = 1
        self.ws_training['F15'] = self.project.training_days
        self.ws_training['F16'] = self.project.training_days

        # 填写G列
        for i in range(4, 18):
            self.ws_training['G{}'.format(i)].number_format = '¥#,##0.00'
            if i in [5, 13]:
                pass
            elif i == 14:
                self.ws_training['G{}'.format(i)] = '=ROUND((SUM(G4,G6:G11))*0.06,2)'
            elif i == 17:
                self.ws_training['G{}'.format(i)] = '=SUM(G4:G16)'
            else:
                self.ws_training['G{}'.format(i)] = '=D{0}*E{0}*F{0}'.format(i)

        # 填充备注
        self.ws_training['A19'] = '注：'
        self.ws_training['B19'] = '（1）100美元='
        self.ws_training['C19'].number_format = '0.00"元人民币"'
        self.ws_training['C19'] = 700
        self.ws_training['C19'].fill = yellow_fill
        self.ws_training['B20'] = '（2）上述费用参照财政部（2008）第2号文举办援外培训班费用开支标准和财务管理办法给定的费用标准报价'

        # 合并需要合并单元格
        self.ws_training.merge_cells('D2:F2')
        self.ws_training.merge_cells('A2:A3')
        self.ws_training.merge_cells('B2:C3')
        self.ws_training.merge_cells('G2:G3')
        self.ws_training.merge_cells('H2:H3')
        self.ws_training.merge_cells('D5:G5')
        self.ws_training.merge_cells('D13:G13')
        self.ws_training.merge_cells('D14:F14')
        self.ws_training.merge_cells('B17:F17')
        self.ws_training.merge_cells('C19:D19')
        self.ws_training.merge_cells('B20:H20')
        for i in range(4, 15):
            self.ws_training.merge_cells('B{0}:C{0}'.format(i))
        self.ws_training.merge_cells('B15:B16')
        self.ws_training.merge_cells('A15:A16')

project = Project('project.docx')
# project.show_info()
my_quota = Quotation(project)
my_quota.create_all('投标报价表-{}.xlsx'.format(my_quota.project.name))