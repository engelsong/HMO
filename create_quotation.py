#!/usr/bin/python3
# encoding: utf-8
"""
@version: python3.6
@author: ‘song‘
@software: PyCharm
@file: examination.py.py
@time: 2018/4/12 22:00
"""
import cmath
from os import linesep, popen
from datetime import datetime
from docx import Document
from openpyxl import Workbook
from openpyxl.styles import Border, Side, Alignment, Font, PatternFill
from openpyxl.workbook.properties import CalcProperties
from openpyxl.worksheet.page import PageMargins
from openpyxl.styles.differential import DifferentialStyle
from openpyxl.formatting.rule import CellIsRule


class Project(object):
    """通过Word文档建立项目对象保存项目信息"""

    def __init__(self, document_name):
        self.name = None
        self.code = None
        self.date = None
        self.destination = None
        self.trans = None
        self.totalsum = 0
        self.is_lowprice = False  # 是否为低价法
        self.is_tech = False  # 是否有技术服务
        self.is_qa = False  # 是否有售后
        self.is_cc = False  # 是否来华培训
        self.techinfo = []  # 存放技术服务信息，格式为[人数，天数，[伙食费，住宿费，公杂费]]
        self.training_days = 0  # 来华培训天数
        self.training_num = 0  # 来华培训人数
        self.qc = []  # 法检物资序号
        self.commodities = {}  # 存放物资信息字典
        document = Document(document_name)
        table1, table2 = document.tables  # 读取两个表格
        project_info = []
        for cell in table1.column_cells(1):
            project_info.append(cell.text)
        table2_length = len(table2.rows)
        for index in range(1, table2_length):  # 从第2行开始读取表格
            temp = []
            row_now = table2.row_cells(index)
            length_row = len(row_now)
            for i in range(1, length_row):  # 将每行信息放入暂存数组
                temp.append(row_now[i].text)
            temp.append(row_now[0].text)  # 把物资编号放在最后一位
            self.commodities[index] = temp
        self.name, self.code, self.date, self.destination, self.trans = project_info[0:5]
        self.totalsum = int(project_info[5])
        if project_info[6] in 'yY':
            self.is_lowprice = True
        if project_info[7] in 'yY':
            self.is_tech = True
            self.techinfo += list(map(int, project_info[8:10]))
            self.techinfo.append(list(map(int, project_info[10].split())))
        if project_info[11] in 'yY':
            self.is_qa = True
        if project_info[12] in 'yY':
            self.is_cc = True
            self.training_days = int(project_info[14])  # 读取来华陪训天数
            self.training_num = int(project_info[13])  # 读取来华培训人数
        if project_info[-1] != '':
            self.qc += list(map(int, project_info[-1].split()))
            self.qc.sort()

    def show_info(self):
        print('项目名称:', self.name)
        print('项目代码:', self.code)
        print('开标日期:', self.date)
        print('目的地:', self.destination)
        print('运输方式:', self.trans)
        print('对外货值：', self.totalsum)
        print('是否为低价法', '是' if self.is_lowprice is True else '否')
        print('是否有技术服务:', '是' if self.is_tech is True else '否')
        print('是否有售后服务:', '是' if self.is_qa is True else '否')
        print('是否有来华培训', '是' if self.is_cc is True else '否')
        if self.is_tech:
            print('技术服务人数:', self.techinfo[0])
            print('技术服务天数:', self.techinfo[1])
            print('伙食费:', self.techinfo[2][0])
            print('住宿费:', self.techinfo[2][1])
            print('公杂费:', self.techinfo[2][2])
        if self.is_cc:
            print('来华培训人数：', self.training_num)
            print('来华培训天数：', self.training_days)
        if len(self.qc) > 0:
            print('法检物资：', self.qc)

    def show_commoditiy(self):
        temp_list = sorted(list(self.commodities.keys()))
        for i in temp_list:
            print(i)
            for j in self.commodities[i]:
                print(j)


class Quotation(object):
    """通过project实例创建报价表"""

    def __init__(self, project):
        self.project = project
        self.wb = Workbook()
        # self.wb.calculation = CalcProperties(iterate=True)
        self.ws_input = None
        self.ws_cost = None
        self.ws_examination = None
        self.ws_lawexam = None
        self.ws_techserve = None
        self.ws_selection = None
        self.ws_itemized_quotation = None
        self.ws_summed_quotation = None
        self.ws_general = None
        self.col_name = ''

    def create_all(self):
        self.create_input()
        self.create_cost()
        self.create_selection()
        self.create_examination()
        if len(self.project.qc) > 0:
            self.create_lawexam()
        if self.project.is_cc:
            self.create_training()
        if self.project.is_tech:
            self.create_techserve()
        self.create_itemized_quotation()
        self.create_summed_quotation()
        self.create_general()
        self.wb.calculation = CalcProperties(iterate=True)
        self.wb.save('投标报价表-{}.xlsx'.format(self.project.name))

    def create_general(self):
        """创建总报价表"""
        self.ws_general = self.wb.create_sheet('1.报价总表', 3)
        colum_title = ['序号', '费用项目', '合计金额', '备注']
        title_width = [10, 35, 25, 20]
        colum_number = len(colum_title)
        row_number = 7

        # 设置基本的样式
        real_side = Side(style='thin')
        full_border = Border(
            left=real_side,
            right=real_side,
            top=real_side,
            bottom=real_side)
        ctr_alignment = Alignment(
            horizontal='center',
            vertical='center',
            wrap_text=True)
        bold_font = Font(name='宋体', bold=True, size=14)
        normal_font = Font(name='宋体', size=14)
        title_font = Font(name='黑体', bold=True, size=20)

        # 合并需要合并单元格
        self.ws_general.merge_cells('A1:D1')

        # 初始化表格
        for i in range(colum_number):
            for j in range(1, row_number):
                cell_now = self.ws_general.cell(row=j + 1, column=i + 1)
                cell_now.border = full_border
                if j < 2:
                    cell_now.font = bold_font
                else:
                    cell_now.font = normal_font
                if i == 1 and j > 1:
                    cell_now.alignment = Alignment(
                        horizontal='left', vertical='center', wrap_text=True)
                elif i == 2 and j > 1:
                    cell_now.alignment = Alignment(
                        horizontal='right', vertical='center', wrap_text=True)
                    cell_now.number_format = '¥#,##0.00'
                else:
                    cell_now.alignment = ctr_alignment
        for i in range(len(title_width)):  # 修改列宽
            self.ws_general.column_dimensions[
                self.ws_general.cell(row=4, column=i + 1).column_letter].width = title_width[i]
        self.ws_general.row_dimensions[2].height = 40

        # 创建标题行
        self.ws_general.merge_cells('A1:D1')
        self.ws_general['A1'].font = title_font
        self.ws_general['A1'].alignment = ctr_alignment
        self.ws_general['A1'] = '1.报价总表'
        self.ws_general.row_dimensions[1].height = 50

        # 填写表头
        index = 0
        for i in self.ws_general['A2':'D2'][0]:
            # print(index+1, i)
            if colum_title[index] != '':
                i.value = colum_title[index]
                i.font = bold_font
            index += 1

        # 填写数据
        self.ws_general['A3'] = '一'
        self.ws_general['B3'] = "全部物资价格（{}{}）{}含商品进价成本、流通环节费用、资金占用成本、合理利润、税金".format(
            self.project.trans, self.project.destination, linesep)
        self.ws_general['C3'] = "='3.物资对内分项报价表'!M{}".format(
            len(self.project.commodities) + 3)
        self.ws_general['A4'] = "二"
        real_row_num = 5  # 判断行数
        if self.project.is_tech:
            real_row_num += 1
            self.ws_general['C4'] = "='4.技术服务费报价表'!H16"
            self.ws_general['B4'] = '技术服务费'
            if self.project.is_cc:
                self.ws_general['C5'] = "='5.来华培训费报价表'!G17"
                self.ws_general['B5'] = '来华培训费'
                real_row_num += 1
        elif self.project.is_cc:
            self.ws_general['C4'] = "='5.来华培训费报价表'!G17"
            self.ws_general['B4'] = '来华培训费'
            real_row_num += 1
        no_seq = ['二', '三', '四']
        for i in range(4, real_row_num):
            self.ws_general['A{}'.format(i)] = no_seq[i - 4]
        self.ws_general["B{}".format(real_row_num - 1)] = "其他费用{}含须中方承担的其他费用、管理费、风险预涨费、" \
                                                          "防恐措施费（如有）".format(
                                                              linesep)
        self.ws_general['C{}'.format(real_row_num - 1)] = "=费用输入!J14"
        self.ws_general['B{}'.format(real_row_num)] = '合计'
        self.ws_general['C{}'.format(real_row_num)] = "=SUM(C3:C{})".format(
            real_row_num - 1)
        self.ws_general['C{}'.format(real_row_num)].font = bold_font
        for i in range(7, real_row_num, -1):
            self.ws_general.delete_rows(i)

        # 打印设置
        self.ws_general.print_options.horizontalCentered = True
        self.ws_general.print_area = 'A1:D{}'.format(real_row_num)
        self.ws_general.page_setup.fitToWidth = 1
        # self.ws_general.page_margins = Content.margin

    def create_input(self):
        """创建物资输入表"""
        self.ws_input = self.wb.create_sheet('物资输入', 0)
        colum_title = ['序号', '品名', 'HS编码', '数量', '', '品牌', '规格、技术参数或配置', '单价',
                       '总价', '生产厂商', '供货商', '产地', '生产或供货地', '供货联系人',
                       '联系电话', '出厂日期', '出口港', '检验标准', '产地或供货地检验（查验）机构', '装运前核验机构',
                       '口岸检装机构', '交货期', '交货地点', '件数说明', '备注']
        title_width = [6, 14, 12, 3, 5, 10, 30, 14, 16, 10, 10, 10, 15, 14, 14, 10,
                       10, 16, 16, 16, 16, 10, 10, 10, 6]

        # 设置基本的样式
        real_side = Side(style='thin')
        full_border = Border(
            left=real_side,
            right=real_side,
            top=real_side,
            bottom=real_side)
        ctr_alignment = Alignment(
            horizontal='center',
            vertical='center',
            wrap_text=True)
        bold_font = Font(name='宋体', bold=True, size=12)
        normal_font = Font(name='宋体', size=10)

        # 初始化表格
        colum_number = len(colum_title)
        row_number = len(self.project.commodities) + 1
        for i in range(colum_number):
            for j in range(row_number):
                cell_now = self.ws_input.cell(row=j + 1, column=i + 1)
                cell_now.border = full_border
                cell_now.font = normal_font
                cell_now.alignment = ctr_alignment

        for i in range(len(title_width)):  # 修改列宽
            self.ws_input.column_dimensions[
                self.ws_input.cell(row=1, column=i + 1).column_letter].width = title_width[i]

        # 填写表头
        index = 0
        for i in self.ws_input['A1':'Y1'][0]:
            i.value = colum_title[index]
            i.font = bold_font
            i.alignment = ctr_alignment
            index += 1

        # 填写物资数据
        relate_coord = [('B', 0), ('C', 1), ('D', 2), ('R', 5)]
        for num in range(2, row_number + 1):
            if self.project.commodities[num - 1][-1] == '':
                self.ws_input['A{}'.format(num)] = num - 1
            else:
                self.ws_input['A{}'.format(
                    num)] = self.project.commodities[num - 1][-1]  # 填写物资序号
            self.ws_input['H{}'.format(num)].number_format = '¥#,##0.00'
            # self.ws_input['H{}'.format(num)].value = 1
            self.ws_input['I{}'.format(num)].number_format = '¥#,##0.00'
            self.ws_input['Y{}'.format(num)] = '-'
            for rela in relate_coord:
                self.ws_input['{}{}'.format(
                    rela[0], num)] = self.project.commodities[num - 1][rela[1]]
            else:
                self.ws_input['E{}'.format(num)].number_format = '0'
                self.ws_input['E{}'.format(num)] = int(
                    self.project.commodities[num - 1][3])
            self.ws_input['I{}'.format(num)] = '=E{}*H{}'.format(num, num)
        # self.wb.save('sample.xlsx')

        self.ws_input.merge_cells('D1:E1')

    def create_cost(self):
        """创建费用输入表格"""
        self.ws_cost = self.wb.create_sheet('费用输入', 1)
        colum_title = [
            '海运',
            '单价',
            '',
            '数量',
            '总金额',
            '陆运',
            '单价',
            '',
            '数量',
            '总金额']
        title_width = [14, 12, 12, 8, 16, 14, 16, 12, 8, 16]

        # 设置基本的样式
        real_side = Side(style='thin')
        full_border = Border(
            left=real_side,
            right=real_side,
            top=real_side,
            bottom=real_side)
        ctr_alignment = Alignment(
            horizontal='center',
            vertical='center',
            wrap_text=True)
        bold_font = Font(name='宋体', bold=True, size=12)
        normal_font = Font(name='宋体', size=12)
        yellow_fill = PatternFill(
            fill_type='solid',
            start_color='FFFF00',
            end_color='FFFF00')

        # 初始化表格
        colum_number = len(colum_title)
        row_number = 18
        for i in range(colum_number):
            for j in range(row_number):
                cell_now = self.ws_cost.cell(row=j + 1, column=i + 1)
                cell_now.border = full_border
                cell_now.font = normal_font
                cell_now.alignment = ctr_alignment
        for i in range(len(title_width)):  # 修改列宽
            self.ws_cost.column_dimensions[
                self.ws_cost.cell(row=4, column=i + 1).column_letter].width = title_width[i]
        for i in range(1, row_number + 1):  # 修改行高
            self.ws_cost.row_dimensions[
                self.ws_cost.cell(row=i, column=1).row].height = 24

        # 填写表头
        index = 0
        for i in self.ws_cost['A1':'J1'][0]:
            # print(index+1, i)
            if colum_title[index] != '':
                i.value = colum_title[index]
                i.font = bold_font
            index += 1

        # 按列从左到右填写表格
        self.ws_cost['A2'] = '全程运费'
        self.ws_cost['A5'] = '港杂费'
        self.ws_cost['A8'] = '仓库装箱费'
        self.ws_cost['A11'] = '运抵报告费'
        self.ws_cost['A14'] = '加固费'
        self.ws_cost['A15'] = '舱单费'
        self.ws_cost['A16'] = '文件费'
        self.ws_cost['A17'] = '报关费'
        self.ws_cost['A18'] = '苫布费'
        index = 2
        for i in ['20GP', '40GP/HQ', '40FR'] * 4:
            self.ws_cost['B{}'.format(index)] = i
            index += 1
        for i in range(2, 14):
            if i < 5:  # 美元部分格式及E列公式
                self.ws_cost['C{}'.format(i)].number_format = '$#,##0.00'
                self.ws_cost['E{}'.format(i)].number_format = '$#,##0.00'
                self.ws_cost['H{}'.format(i)].number_format = '$#,##0.00'
                self.ws_cost['E{}'.format(i)] = '=C{0}*D{0}*F10'.format(i)
            else:
                self.ws_cost['C{}'.format(i)].number_format = '¥#,##0.00'
            self.ws_cost['C{}'.format(i)] = 0
        self.ws_cost['C14'] = 1000
        self.ws_cost['C14'].number_format = '0"元/箱"'
        self.ws_cost['C15'] = 100
        self.ws_cost['C15'].number_format = '0"元/票"'
        self.ws_cost['C16'] = 500
        self.ws_cost['C16'].number_format = '0"元/票"'
        self.ws_cost['C17'] = 300
        self.ws_cost['C17'].number_format = '0"元/票"'
        self.ws_cost['C18'] = 0
        self.ws_cost['C18'].number_format = '0"元/柜"'
        for i in range(2, 19):  # D列格式及E列格式
            self.ws_cost['D{}'.format(i)] = 0
            self.ws_cost['D{}'.format(i)].number_format = '0'
            self.ws_cost['E{}'.format(i)].number_format = '¥#,##0.00'
            if i > 4:
                self.ws_cost['E{}'.format(
                    i)] = '=C{0}*D{0}'.format(i)  # E 列公式生成
                self.ws_cost['D{}'.format(i)].number_format = '0'
        self.ws_cost['F2'] = '内陆运费'
        self.ws_cost['F5'] = '电子跟踪单'
        self.ws_cost['F7'] = '货物为车辆'
        self.ws_cost['F9'] = '运费合计'
        self.ws_cost['F9'].font = bold_font
        self.ws_cost['F10'].font = bold_font
        self.ws_cost['F10'].fill = yellow_fill
        self.ws_cost['F10'] = 6.9
        self.ws_cost['F10'].number_format = '"汇率："0.00'
        self.ws_cost['F15'] = '商检费用'
        self.ws_cost['F15'].font = bold_font
        self.ws_cost['F17'] = '对外货值'
        self.ws_cost['F18'] = '保险费用'
        self.ws_cost['F18'].font = bold_font
        self.ws_cost['G2'] = '20GP'
        self.ws_cost['G3'] = '40GP/HQ'
        self.ws_cost['G4'] = '40FR'
        self.ws_cost['G5'] = '不超过5个'
        self.ws_cost['G6'] = '超过5个追加'
        self.ws_cost['G7'] = '不超过5个'
        self.ws_cost['G8'] = '超过5个追加'
        self.ws_cost['G17'] = self.project.totalsum
        self.ws_cost['G17'].number_format = '¥#,##0.00'
        for i in range(2, 9):
            self.ws_cost['H{}'.format(i)] = 0
            self.ws_cost['I{}'.format(i)] = 0
            self.ws_cost['J{}'.format(i)] = '=H{0}*I{0}*F10'.format(i)
            self.ws_cost['J{}'.format(i)].number_format = '¥#,##0.00'
        self.ws_cost['J9'] = '=SUM(E2:E18)+SUM(J2:J8)'
        self.ws_cost['J9'].font = bold_font
        self.ws_cost['J9'].number_format = '¥#,##0.00'
        self.ws_cost['H17'] = '费率'
        self.ws_cost['I17'] = 0.001
        self.ws_cost['I17'].number_format = '0.00%'
        self.ws_cost['I17'].fill = yellow_fill
        self.ws_cost['J15'].number_format = '¥#,##0.00'
        self.ws_cost['J15'].fill = yellow_fill
        self.ws_cost['J18'].font = bold_font
        self.ws_cost['J18'].number_format = '¥#,##0.00'
        self.ws_cost['J18'] = '=round(G17*1.1*I17,2)'
        self.ws_cost['F14'] = '管理费'
        self.ws_cost['F14'].font = bold_font
        self.ws_cost['J14'].number_format = '¥#,##0.00'
        self.ws_cost['J14'] = 0
        self.ws_cost['J14'].fill = yellow_fill
        self.ws_cost['J15'] = 0

        # 合并需要合并单元格
        self.ws_cost.merge_cells('B1:C1')
        self.ws_cost.merge_cells('G1:H1')
        self.ws_cost.merge_cells('F9:I9')
        self.ws_cost.merge_cells('F10:J10')
        self.ws_cost.merge_cells('F14:I14')
        self.ws_cost.merge_cells('F15:I15')
        self.ws_cost.merge_cells('F18:I18')
        self.ws_cost.merge_cells('F2:F4')
        self.ws_cost.merge_cells('F5:F6')
        self.ws_cost.merge_cells('F7:F8')
        for i in range(2, 14, 3):
            self.ws_cost.merge_cells('A{}:A{}'.format(i, i + 2))

    def create_selection(self):
        """创建物资选型一览表"""
        self.ws_selection = self.wb.create_sheet('0.物资选型一览表', 2)
        colum_title = ['序号', '物资名称', '招标要求', '投标型号及规格', '响应/偏离', '生产企业', '交货期',
                       '交货地点', '说明']
        title_width = [6, 12, 50, 75, 8, 6, 6, 6, 6]

        # 设置基本的样式
        real_side = Side(style='thin')
        full_border = Border(
            left=real_side,
            right=real_side,
            top=real_side,
            bottom=real_side)
        ctr_alignment = Alignment(
            horizontal='center',
            vertical='center',
            wrap_text=True)
        left_alignment = Alignment(
            horizontal='left',
            vertical='top',
            wrap_text=True)
        bold_font = Font(name='宋体', bold=True, size=12)
        normal_font = Font(name='宋体', size=10)
        title_font = Font(name='黑体', size=20)

        # 初始化表格
        colum_number = len(colum_title)
        row_number = len(self.project.commodities) + 2
        for i in range(colum_number):
            for j in range(row_number):
                cell_now = self.ws_selection.cell(row=j + 1, column=i + 1)
                if j > 0:  # 第一列留下给表头
                    cell_now.border = full_border
                    cell_now.font = normal_font
                    cell_now.alignment = ctr_alignment
        for i in range(len(title_width)):  # 修改列宽
            self.ws_selection.column_dimensions[
                self.ws_selection.cell(row=1, column=i + 1).column_letter].width = title_width[i]

        # 创建标题行
        self.ws_selection.merge_cells('A1:I1')
        self.ws_selection['A1'].font = title_font
        self.ws_selection['A1'].alignment = ctr_alignment
        self.ws_selection['A1'] = '2.物资选型一览表'
        self.ws_selection.row_dimensions[1].height = 40

        # 填写表头
        index = 0
        for i in self.ws_selection['A2':'I2'][0]:
            i.value = colum_title[index]
            i.font = bold_font
            i.alignment = ctr_alignment
            index += 1

        # 填入数据
        col_relate = [('A', 'A'), ('B', 'B'), ('D', 'G'),
                      ('F', 'J'), ('G', 'V'), ('H', 'W')]
        for row in range(3, row_number + 1):  # 遍历行
            for col in col_relate:  # 根据对应关系设立公式
                cell_now = self.ws_selection['{}{}'.format(col[0], row)]
                cell_now.value = '=物资输入!{}{}'.format(col[1], row - 1)
                if col[0] == 'D':
                    cell_now.alignment = left_alignment
            self.ws_selection['C{}'.format(
                row)] = self.project.commodities[row - 2][4]  # 填入招标要求
            self.ws_selection['C{}'.format(row)].alignment = left_alignment
            self.ws_selection['E{}'.format(row)] = '响应'

        # 打印设置
        self.ws_selection.print_options.horizontalCentered = True
        self.ws_selection.print_area = 'A1:I{}'.format(row_number)
        self.ws_selection.page_setup.fitToWidth = 1
        self.ws_selection.page_setup.orientation = "landscape"
        self.ws_selection.page_margins = PageMargins(left=0.25, right=0.25, top=0.75, bottom=0.75, header=0.3,
                                                     footer=0.3)

    def create_itemized_quotation(self):
        """生成分项报价表垂直方向"""
        self.ws_itemized_quotation = self.wb.create_sheet('3.物资对内分项报价表', 3)
        colum_title = ['序号', '费用科目', '1.商品进价成本', '2.国内运杂费', '3.包装费', '4.保管费', '5.物资检验费',
                       '6.保险费', '7.国外运费', '8.资金占用成本', '9.合理利润', '10.税金', '合计']
        title_width = [6, 15, 14, 10, 10, 10, 10, 10, 14, 12, 12, 12, 16]
        colum_number = len(colum_title)
        row_number = len(self.project.commodities) + 6

        # 设置基本的样式
        real_side = Side(style='thin')
        full_border = Border(
            left=real_side,
            right=real_side,
            top=real_side,
            bottom=real_side)
        ctr_alignment = Alignment(
            horizontal='center',
            vertical='center',
            wrap_text=True)
        bold_font = Font(name='宋体', bold=True, size=12)
        normal_font = Font(name='宋体', size=12)
        title_font = Font(name='黑体', size=14)
        right_alignment = Alignment(
            horizontal='right',
            vertical='center',
            wrap_text=False)
        left_alignment = Alignment(
            horizontal='left',
            vertical='center',
            wrap_text=False)
        yellow_fill = PatternFill(
            fill_type='solid',
            start_color='FFFF00',
            end_color='FFFF00')

        # 初始化表格
        for i in range(colum_number):
            for j in range(1, row_number):  # 留出第一行
                cell_now = self.ws_itemized_quotation.cell(
                    row=j + 1, column=i + 1)
                if j < 2:
                    cell_now.font = bold_font
                    cell_now.alignment = ctr_alignment
                else:
                    cell_now.font = normal_font
                if i > 1 and row_number - 2 > j > 1:  # 格式化单元格
                    cell_now.number_format = '#,##0.00'
                    cell_now.alignment = right_alignment
                else:
                    cell_now.alignment = ctr_alignment
                if j < row_number - 3:
                    cell_now.border = full_border
                else:
                    cell_now.alignment = left_alignment
        for i in range(len(title_width)):  # 修改列宽
            self.ws_itemized_quotation.column_dimensions[
                self.ws_itemized_quotation.cell(row=4, column=i + 1).column_letter].width = title_width[i]
        for i in range(3, row_number + 1):  # 修改行高
            self.ws_itemized_quotation.row_dimensions[i].height = 20
        self.ws_itemized_quotation.row_dimensions[2].height = 32

        # 创建标题行
        self.ws_itemized_quotation['A1'].font = title_font
        self.ws_itemized_quotation['A1'].alignment = ctr_alignment
        self.ws_itemized_quotation['A1'] = '3.物资对内分项报价表'
        self.ws_itemized_quotation.row_dimensions[1].height = 30

        # 填写表头
        index = 0
        for i in self.ws_itemized_quotation['A2':'M2'][0]:
            i.value = colum_title[index]
            i.font = bold_font
            i.alignment = ctr_alignment
            index += 1

        # 填写数据
        self.ws_itemized_quotation['A{}'.format(row_number - 2)] = '注：'
        self.ws_itemized_quotation['B{}'.format(row_number - 2)] = '1.第8项资金占用成本=（商品进价成本+物资检验费+保险费' \
                                                                   '+国外运费）×4.35%利率×预计占用3个月/12个月'
        self.ws_itemized_quotation['B{}'.format(row_number - 1)] = '2.第9项合理利润=200万×4%+300万×3.5%+500万×3%+1000万' \
                                                                   '×2%+3000万×1%+（X-5000万）×0.75%'
        self.ws_itemized_quotation['B{}'.format(
            row_number)] = '3.第10项应纳增值税=[对内总承包价/（1+增值税税率）]X增值税税率-当期进项税款'
        self.ws_itemized_quotation['B{}'.format(
            row_number - 2)].fill = yellow_fill
        self.ws_itemized_quotation['B{}'.format(
            row_number - 1)].fill = yellow_fill
        self.ws_itemized_quotation['B{}'.format(row_number - 3)] = '小计'

        col_relate = [('A', 'A'), ('B', 'B'), ('C', 'I')]
        for row in range(3, row_number - 3):
            for col in col_relate:  # 根据对应关系设立公式
                self.ws_itemized_quotation['{}{}'.format(
                    col[0], row)] = '=物资输入!{}{}'.format(col[1], row - 1)
            self.ws_itemized_quotation['D{}'.format(row)] = 0
            self.ws_itemized_quotation['E{}'.format(row)] = 0
            self.ws_itemized_quotation['F{}'.format(row)] = 0
            self.ws_itemized_quotation['G{}'.format(
                row)] = '=round(C{0}/C{1}*G{1},2)'.format(row, row_number - 3)
            self.ws_itemized_quotation['H{}'.format(
                row)] = '=round(C{0}/C{1}*H{1},2)'.format(row, row_number - 3)
            self.ws_itemized_quotation['I{}'.format(
                row)] = '=round(C{0}/C{1}*I{1},2)'.format(row, row_number - 3)
            self.ws_itemized_quotation['J{}'.format(
                row)] = '=round(C{0}/C{1}*J{1},2)'.format(row, row_number - 3)
            self.ws_itemized_quotation['K{}'.format(
                row)] = '=round(C{0}/C{1}*K{1},2)'.format(row, row_number - 3)
            self.ws_itemized_quotation['L{}'.format(
                row)] = '=round(C{0}/C{1}*L{1},2)'.format(row, row_number - 3)
            self.ws_itemized_quotation['M{}'.format(
                row)] = '=SUM(C{0}:L{0})'.format(row)
        self.ws_itemized_quotation['C{}'.format(
            row_number - 3)] = '=SUM(C3:C{})'.format(row_number - 4)
        self.ws_itemized_quotation['D{}'.format(
            row_number - 3)] = '=SUM(D3:D{})'.format(row_number - 4)
        self.ws_itemized_quotation['E{}'.format(
            row_number - 3)] = '=SUM(E3:E{})'.format(row_number - 4)
        self.ws_itemized_quotation['F{}'.format(
            row_number - 3)] = '=SUM(F3:F{})'.format(row_number - 4)
        self.ws_itemized_quotation['G{}'.format(row_number - 3)] = '=费用输入!J15'
        self.ws_itemized_quotation['H{}'.format(row_number - 3)] = '=费用输入!J18'
        self.ws_itemized_quotation['I{}'.format(row_number - 3)] = '=费用输入!J9'
        self.ws_itemized_quotation['K{}'.format(
            row_number - 3)] = '=round(IF(C{0}>50000000,(C{0}-50000000)*0.0075+835000,IF(C{0}>20000000,(C{0}-20000000' \
                               ')*0.01+535000,IF(C{0}>10000000,(C{0}-10000000)*0.02+335000,IF(C{0}>5000000,(C{0}-' \
                               '5000000)*0.03+185000,IF(C{0}>2000000,(C{0}-2000000)*0.035+80000,C{0}*0.04))))),2)'\
            .format(row_number - 3)
        self.ws_itemized_quotation['L{}'.format(row_number - 3)] = \
            '=round((M{0}/1.13*0.13-C{0}/1.13*0.13-G{0}/1.06*0.06),2)'.format(
                row_number - 3)
        self.ws_itemized_quotation['M{}'.format(
            row_number - 3)] = '=SUM(C{0}:L{0})'.format(row_number - 3)
        self.ws_itemized_quotation['J{}'.format(
            row_number - 3)] = '=round(SUM(C{0}:I{0})*3/12*0.0435,2)'.format(row_number - 3)
        self.ws_itemized_quotation['J{}'.format(
            row_number - 3)].fill = yellow_fill
        self.ws_itemized_quotation['N{}'.format(
            row_number - 3)] = '=SUM(M3:M{})'.format(row_number - 4)

        # 低价项目针对部分单元格进行修改
        if self.project.is_lowprice:
            for row in range(3, row_number - 3):
                self.ws_itemized_quotation['G{}'.format(row)] = 0.01
                self.ws_itemized_quotation['H{}'.format(row)] = 0.01
                self.ws_itemized_quotation['I{}'.format(row)] = 0.01
            self.ws_itemized_quotation['G{}'.format(row_number - 3)] = '=sum(G3:G{})'.format(row_number - 4)
            self.ws_itemized_quotation['H{}'.format(row_number - 3)] = '=sum(H3:H{})'.format(row_number - 4)
            self.ws_itemized_quotation['I{}'.format(row_number - 3)] = '=sum(I3:I{})'.format(row_number - 4)
            self.ws_itemized_quotation['J{}'.format(
                row_number - 3)] = '=round((SUM(C{0}:I{0})*3/12*0.0435)*0.8,2)'.format(row_number - 3)
            self.ws_itemized_quotation['K{}'.format(
                row_number - 3)] = '=round(IF(C{0}>50000000,(C{0}-50000000)*0.0075+835000,IF(C{0}>20000000,' \
                                   '(C{0}-20000000)*0.01+535000,IF(C{0}>10000000,(C{0}-10000000)*0.02+335000,' \
                                   'IF(C{0}>5000000,(C{0}-5000000)*0.03+185000,IF(C{0}>2000000,' \
                                   '(C{0}-2000000)*0.035+80000,C{0}*0.04)))))*0.8,2)'.format(row_number - 3)
            self.ws_itemized_quotation['L{}'.format(row_number - 3)] = \
                '=round((M{0}/1.13*0.13-C{0}/1.13*0.13-G{0}/1.06*0.06)*0.9,2)'.format(
                    row_number - 3)

        # 增加条件格式判断
        red_fill = PatternFill(
            start_color='EE1111',
            end_color='EE1111',
            fill_type='solid')
        self.ws_itemized_quotation.conditional_formatting.add('N{}'.format(row_number - 3), CellIsRule(
            operator='notEqual', formula=['M{}'.format(row_number - 3)], fill=red_fill))

        # 合并需要合并单元格
        self.ws_itemized_quotation.merge_cells('A1:M1')
        self.ws_itemized_quotation.merge_cells(
            'B{0}:M{0}'.format(row_number - 1))
        self.ws_itemized_quotation.merge_cells(
            'B{0}:M{0}'.format(row_number - 2))
        self.ws_itemized_quotation.merge_cells('B{0}:M{0}'.format(row_number))

        # 打印设置
        self.ws_itemized_quotation.print_options.horizontalCentered = True
        self.ws_itemized_quotation.print_area = 'A1:M{}'.format(row_number)
        self.ws_itemized_quotation.page_setup.fitToWidth = 1
        self.ws_itemized_quotation.page_setup.orientation = "landscape"
        self.ws_itemized_quotation.page_margins = PageMargins(left=0.25, right=0.25, top=0.75, bottom=0.75, header=0.3,
                                                              footer=0.3)

    def create_summed_quotation(self):
        """创建对内总报价表"""
        self.ws_summed_quotation = self.wb.create_sheet('2.物资对内总报价表', 3)
        colum_title = ['序号', '品名', '规格', '商标', '产地', '生产年份', '{}单价{}(元人民币)'.format(
            self.project.trans, linesep), '数量', '', '{}总价{}(元人民币)'.format(self.project.trans, linesep), '备注']
        title_width = [6, 8, 50, 6, 6, 6, 14, 6, 6, 14, 10]
        colum_number = len(colum_title)
        row_number = len(self.project.commodities) + 4

        # 设置基本的样式
        real_side = Side(style='thin')
        full_border = Border(
            left=real_side,
            right=real_side,
            top=real_side,
            bottom=real_side)
        ctr_alignment = Alignment(
            horizontal='center',
            vertical='center',
            wrap_text=True)
        bold_font = Font(name='宋体', bold=True, size=12)
        normal_font = Font(name='宋体', size=12)
        title_font = Font(name='黑体', size=20)
        left_alignment = Alignment(
            horizontal='left',
            vertical='top',
            wrap_text=True)

        # 初始化表格
        for i in range(colum_number):
            for j in range(1, row_number):
                cell_now = self.ws_summed_quotation.cell(
                    row=j + 1, column=i + 1)
                if j < row_number - 1:
                    cell_now.border = full_border
                if j < 2:
                    cell_now.font = bold_font
                else:
                    cell_now.font = normal_font
                cell_now.alignment = ctr_alignment
        for i in range(len(title_width)):  # 修改列宽
            self.ws_summed_quotation.column_dimensions[
                self.ws_summed_quotation.cell(row=4, column=i + 1).column_letter].width = title_width[i]
        for i in range(2, row_number):  # 修改行高
            self.ws_summed_quotation.row_dimensions[
                self.ws_summed_quotation.cell(row=i, column=1).row].height = 30
        self.ws_summed_quotation.row_dimensions[row_number].height = 45

        # 创建标题行
        self.ws_summed_quotation['A1'].font = title_font
        self.ws_summed_quotation['A1'].alignment = ctr_alignment
        self.ws_summed_quotation['A1'] = '2.物资对内总报价表'
        self.ws_summed_quotation.row_dimensions[1].height = 30

        # 填写表头
        index = 0
        for i in self.ws_summed_quotation['A2':'k2'][0]:
            # print(index+1, i)
            if colum_title[index] != '':
                i.value = colum_title[index]
                # i.font = bold_font
            index += 1

        # 填入数据
        col_relate = [('A', 'A'), ('B', 'B'), ('C', 'G'), ('D', 'F'), ('E', 'L'), ('F', 'P'), ('H', 'D'),
                      ('I', 'E'), ('K', 'X')]
        for row in range(3, row_number - 1):  # 遍历行
            for col in col_relate:  # 根据对应关系设立公式
                cell_now = self.ws_summed_quotation['{}{}'.format(col[0], row)]
                cell_now.value = '=物资输入!{}{}'.format(col[1], row - 1)
                if col[0] == 'C':
                    cell_now.alignment = left_alignment
            self.ws_summed_quotation['J{}'.format(
                row)] = "='3.物资对内分项报价表'!M{}".format(row)
            self.ws_summed_quotation['J{}'.format(
                row)].number_format = '#,##0.00'
            self.ws_summed_quotation['G{}'.format(
                row)] = '=round(J{0}/I{0},2)'.format(row)
            self.ws_summed_quotation['G{}'.format(
                row)].number_format = '#,##0.00'
        self.ws_summed_quotation['A{}'.format(row_number - 1)] = '合计金额'
        self.ws_summed_quotation['C{}'.format(
            row_number - 1)] = '=SUM(J3:J{})'.format(row_number - 1)
        self.ws_summed_quotation['C{}'.format(row_number - 1)].font = bold_font
        self.ws_summed_quotation['C{}'.format(
            row_number - 1)].number_format = '¥#,##0.00'
        self.ws_summed_quotation['C{}'.format(row_number - 1)].alignment = Alignment(
            horizontal='left', vertical='center', wrap_text=True)
        self.ws_summed_quotation['A{}'.format(row_number)] = '注：'
        self.ws_summed_quotation['B{}'.format(row_number)] = "（1）本表所列{}{}单价和总价包括投标人提供上述全部物资" \
            "并承担合同规定全部义务所需的一切费用；{}（2）在备注栏中注明包装" \
            "情况,即包装的单位和数量".format(self.project.trans,
                                  self.project.destination, linesep)
        self.ws_summed_quotation['B{}'.format(row_number)].alignment = Alignment(horizontal='left',
                                                                                 vertical='center', wrap_text=True)

        # 合并需要合并单元格
        self.ws_summed_quotation.merge_cells('A1:K1')
        self.ws_summed_quotation.merge_cells('H2:I2')
        self.ws_summed_quotation.merge_cells(
            'A{0}:B{0}'.format(row_number - 1))
        self.ws_summed_quotation.merge_cells(
            'C{0}:K{0}'.format(row_number - 1))
        self.ws_summed_quotation.merge_cells('B{0}:K{0}'.format(row_number))

        # 打印设置
        self.ws_summed_quotation.print_options.horizontalCentered = True
        self.ws_summed_quotation.print_area = 'A1:K{}'.format(row_number)
        self.ws_summed_quotation.page_setup.fitToWidth = 1
        self.ws_summed_quotation.page_setup.orientation = "landscape"
        self.ws_summed_quotation.page_margins = PageMargins(left=0.25, right=0.25, top=0.75, bottom=0.75, header=0.3,
                                                            footer=0.3)

    def create_examination(self):
        """创建物资选型一览表（非法检）"""
        self.ws_examination = self.wb.create_sheet('7.物资检验一览表（非法检物资）', 3)
        colum_title = ['序号', '品名', 'HS编码', '数量及单位', '', '品牌', '规格、技术参数或配置', '金额', '生产厂商',
                       '供货商', '生产或供货地', '供货联系人及联系电话', '', '出厂日期', '出口港', '检验标准', '检验机构名称',
                       '', '', '备注']
        subcol_title = ['产地或供货地检验（查验）机构', '装运前核验机构', '口岸监装机构']
        title_width = [6, 14, 12, 3, 5, 10, 30, 16,
                       16, 16, 10, 10, 10, 10, 8, 18, 13, 8, 7, 6]

        # 设置基本的样式
        real_side = Side(style='thin')
        full_border = Border(
            left=real_side,
            right=real_side,
            top=real_side,
            bottom=real_side)
        ctr_alignment = Alignment(
            horizontal='center',
            vertical='center',
            wrap_text=True)
        left_alignment = Alignment(
            horizontal='left',
            vertical='top',
            wrap_text=True)
        bold_font = Font(name='宋体', bold=True, size=10)
        normal_font = Font(name='宋体', size=10)
        title_font = Font(name='黑体', size=20)

        # 合并需要合并单元格
        self.ws_examination.merge_cells('D2:E3')
        self.ws_examination.merge_cells('Q2:S2')
        self.ws_examination.merge_cells('L2:M3')
        # self.ws_examination['M2'].border = full_border
        for col in range(1, 21):
            if col not in [4, 5, 12, 13, 17, 18, 19]:
                self.ws_examination.merge_cells(
                    start_row=2, start_column=col, end_row=3, end_column=col)
        for i in range(len(title_width)):  # 修改列宽
            self.ws_examination.column_dimensions[
                self.ws_examination.cell(row=4, column=i + 1).column_letter].width = title_width[i]

        # 初始化表格
        colum_number = len(colum_title)
        row_number = len(self.project.commodities) + 3
        for i in range(colum_number):
            for j in range(row_number):
                cell_now = self.ws_examination.cell(row=j + 1, column=i + 1)
                if j > 0:  # 第一列留下给表头
                    cell_now.border = full_border
                    cell_now.font = normal_font
                    cell_now.alignment = ctr_alignment

        # 创建标题行
        self.ws_examination.merge_cells('A1:T1')
        self.ws_examination['A1'].font = title_font
        self.ws_examination['A1'].alignment = ctr_alignment
        index = 4  # 计算表格序号
        if self.project.is_tech:
            index += 1
        if len(self.project.qc) > 0:
            index += 1
        self.ws_examination['A1'] = '{}.物资检验一览表（非法检物资）'.format(index)
        self.ws_examination.row_dimensions[1].height = 30

        # 填写表头
        index = 0
        for i in self.ws_examination['A2':'T2'][0]:
            # print(index+1, i)
            if colum_title[index] != '':
                i.value = colum_title[index]
                i.font = bold_font
            index += 1
        index = 0
        for cell in self.ws_examination['Q3':'S3'][0]:
            cell.value = subcol_title[index]
            cell.font = bold_font
            index += 1

        # 填入数据
        col_relate = [('A', 'A'), ('B', 'B'), ('C', 'C'), ('D', 'D'), ('E', 'E'), ('F', 'F'), ('G', 'G'), ('H', 'I'),
                      ('I', 'J'), ('J', 'K'), ('K', 'M'), ('L', 'N'), ('M',
                                                                       'O'), ('N', 'P'), ('O', 'Q'), ('P', 'R'),
                      ('Q', 'S'), ('R', 'T'), ('S', 'U'), ('T', 'Y')]
        for row in range(4, row_number + 1):  # 遍历行
            for col in col_relate:  # 根据对应关系设立公式
                cell_now = self.ws_examination['{}{}'.format(col[0], row)]
                cell_now.value = '=物资输入!{}{}'.format(col[1], row - 2)
                if col[0] == 'G':
                    cell_now.alignment = left_alignment
                if col[0] == 'H':
                    cell_now.number_format = '¥#,##0.00'

        # 打印设置
        self.ws_examination.print_area = 'A1:T{}'.format(row_number)
        self.ws_examination.page_setup.fitToWidth = 1
        self.ws_examination.page_setup.orientation = "landscape"
        self.ws_examination.page_margins = PageMargins(left=0.25, right=0.25, top=0.75, bottom=0.7, header=0.3,
                                                       footer=0.3)

    def create_techserve(self):
        """创建技术服务费报价表"""
        self.ws_techserve = self.wb.create_sheet('4.技术服务费报价表', 3)
        colum_title = [
            '序号',
            '费用名称',
            '外币单价',
            '人民币单价',
            '人数',
            '天/次数',
            '外币合计',
            '人民币合计']
        title_width = [6, 17, 14, 14, 8, 10, 14, 20]

        # 设置基本的样式
        real_side = Side(style='thin')
        full_border = Border(
            left=real_side,
            right=real_side,
            top=real_side,
            bottom=real_side)
        slash_border = Border(diagonal=real_side, diagonalDown=True, left=real_side, right=real_side,
                              top=real_side, bottom=real_side)
        ctr_alignment = Alignment(
            horizontal='center',
            vertical='center',
            wrap_text=True)
        bold_font = Font(name='宋体', bold=True, size=12)
        normal_font = Font(name='宋体', size=12)
        title_font = Font(name='黑体', size=20)
        yellow_fill = PatternFill(
            fill_type='solid',
            start_color='FFFF00',
            end_color='FFFF00')

        # 合并需要合并单元格
        self.ws_techserve.merge_cells('B9:H9')
        self.ws_techserve.merge_cells('C14:G14')
        self.ws_techserve.merge_cells('C15:F15')
        self.ws_techserve.merge_cells('C16:F16')
        self.ws_techserve.merge_cells('B19:H19')

        # 初始化表格
        colum_number = len(colum_title)
        row_number = 19
        for i in range(colum_number):
            for j in range(1, row_number):
                cell_now = self.ws_techserve.cell(row=j + 1, column=i + 1)
                if j < 16:  # 第一列留下给表头
                    cell_now.border = full_border
                    cell_now.font = normal_font
                    cell_now.alignment = ctr_alignment
                else:
                    cell_now.font = normal_font
                    cell_now.alignment = Alignment(
                        horizontal='left', vertical='center', wrap_text=True)
        for i in range(len(title_width)):  # 修改列宽
            self.ws_techserve.column_dimensions[
                self.ws_techserve.cell(row=4, column=i + 1).column_letter].width = title_width[i]
        for i in range(2, row_number):  # 修改行高
            self.ws_techserve.row_dimensions[
                self.ws_techserve.cell(row=i, column=1).row].height = 20
        self.ws_techserve.row_dimensions[19].height = 40

        # 打上斜线
        cell_coor = ['C3', 'C4', 'C5', 'C14', 'C15', 'C16', 'D6', 'D7', 'D8', 'D10', 'D11', 'D12',
                     'D13', 'F4', 'F5', 'G3', 'G4', 'G5']
        for cell in cell_coor:
            self.ws_techserve[cell].border = slash_border

        # 创建标题行
        self.ws_techserve.merge_cells('A1:H1')
        self.ws_techserve['A1'].font = title_font
        self.ws_techserve['A1'].alignment = ctr_alignment
        self.ws_techserve['A1'] = '4.技术服务费报价表'
        self.ws_techserve.row_dimensions[1].height = 30

        # 填写表头
        index = 0
        for i in self.ws_techserve['A2':'H2'][0]:
            # print(index+1, i)
            if colum_title[index] != '':
                i.value = colum_title[index]
                i.font = bold_font
            index += 1

        # 填写数据
        col_a = [
            1,
            2,
            3,
            4,
            5,
            6,
            7,
            '7-1',
            '7-2',
            '7-3',
            '7-4',
            8,
            9,
            '']  # 序号
        col_b = ['技术服务补贴', '国内差旅费', '国际机票费', '国外伙食费', '国外住宿费', '国外公杂费', '国际旅途中转费',
                 '中转伙食费', '中转住宿费', '中转公杂费', '中转机场费', '管理费', '其他费用', '共计']  # 费用名称
        for index in range(3, 17):
            self.ws_techserve['A{}'.format(index)] = col_a[index - 3]
            self.ws_techserve['B{}'.format(index)] = col_b[index - 3]
        for num in range(3):  # 填写技术服务费单价
            self.ws_techserve['C{}'.format(
                num + 6)].number_format = '$#,##0.00'
            self.ws_techserve['C{}'.format(
                num + 6)] = self.project.techinfo[2][num]
        # 格式化单元格
        for i in [3, 4, 5, 6, 7, 8, 14, 16]:
            self.ws_techserve['H{}'.format(i)].number_format = '¥#,##0.00'
            if i < 9:
                self.ws_techserve['E{}'.format(i)].number_format = '0'
                self.ws_techserve['E{}'.format(i)] = self.project.techinfo[0]
                self.ws_techserve['F{}'.format(i)].number_format = '0'
                if i in [3, 6, 7, 8]:
                    self.ws_techserve['F{}'.format(
                        i)] = self.project.techinfo[1]
            if i < 6:
                self.ws_techserve['D{}'.format(i)].number_format = '¥#,##0.00'
            if i > 5:
                self.ws_techserve['G{}'.format(i)].number_format = '$#,##0.00'
        if self.project.is_lowprice:
            self.ws_techserve['F7'] = self.project.techinfo[1] - 1
        # 填充备注
        self.ws_techserve['A18'] = '注：'
        self.ws_techserve['B18'] = '（1）100美元='
        self.ws_techserve['C18'].number_format = '0.00"元人民币"'
        self.ws_techserve['C18'] = 700
        self.ws_techserve['C18'].fill = yellow_fill

        self.ws_techserve['B19'] = '（2）国外伙食、住宿、工杂费和中转伙食、住宿、工杂费须按照财行[2013]516号文' \
                                   '和财行[2017]434号文规定的该国（地区）费用标准和币种填报。计算出外币合计后，' \
                                   '应折算出相应人民币合计数'

        # 填充表格
        self.ws_techserve['D3'] = 3000
        if not self.project.is_lowprice:
            self.ws_techserve['D4'] = 2000
            self.ws_techserve['D5'] = 10000
        else:
            self.ws_techserve['D4'] = 0
            self.ws_techserve['D5'] = 0
        self.ws_techserve['D5'].fill = yellow_fill
        for i in range(6, 9):
            self.ws_techserve['G{}'.format(i)] = '=C{0}*E{0}*F{0}'.format(i)
            self.ws_techserve['H{}'.format(i)] = '=G{}*C18/100'.format(i)
        self.ws_techserve['H3'] = '=D3*F3*E3/30'
        self.ws_techserve['H4'] = '=D4*E4'
        self.ws_techserve['H5'] = '=D5*E5'
        self.ws_techserve['H14'] = '=round((H3+H4+H6+H10)*IF(E7<6,0.21,IF(E7<11,0.18,IF(E7<21,0.15,IF(E7<41,0.12,0.09' \
                                   ')))),2)'
        self.ws_techserve['H16'] = '=SUM(H3:H8, H10:H15)'
        self.ws_techserve['G16'] = '=SUM(G6:G8)'

        # 合并需要合并单元格
        self.ws_techserve.merge_cells('B9:H9')
        self.ws_techserve.merge_cells('C14:G14')
        self.ws_techserve.merge_cells('C15:F15')
        self.ws_techserve.merge_cells('C16:F16')
        self.ws_techserve.merge_cells('B19:H19')
        self.ws_techserve.merge_cells('C18:D18')

        # 打印设置
        self.ws_techserve.print_options.horizontalCentered = True
        self.ws_techserve.print_area = 'A1:H{}'.format(row_number)
        self.ws_techserve.page_setup.fitToWidth = 1
        # self.ws_techserve.page_setup.orientation = "landscape"
        self.ws_techserve.page_margins = PageMargins(left=0.7, right=0.7, top=0.75, bottom=0.75, header=0.3,
                                                     footer=0.3)

    def create_lawexam(self):
        """创建物资选型一览表（法检物资）"""
        self.ws_lawexam = self.wb.create_sheet('6.物资检验一览表（法检物资）', 3)
        colum_title = ['序号', '品名', 'HS编码', '数量及单位', '', '品牌', '规格、技术参数或配置', '金额', '生产厂商',
                       '供货商', '生产或供货地', '供货联系人及联系电话', '', '出厂日期', '出口港', '检验标准',
                       '口岸检验机构', '备注']
        title_width = [6, 14, 12, 3, 5, 10, 30, 16,
                       16, 16, 10, 10, 10, 10, 8, 18, 7, 6]

        # 设置基本的样式
        real_side = Side(style='thin')
        full_border = Border(
            left=real_side,
            right=real_side,
            top=real_side,
            bottom=real_side)
        ctr_alignment = Alignment(
            horizontal='center',
            vertical='center',
            wrap_text=True)
        left_alignment = Alignment(
            horizontal='left',
            vertical='top',
            wrap_text=True)
        bold_font = Font(name='宋体', bold=True, size=10)
        normal_font = Font(name='宋体', size=10)
        title_font = Font(name='黑体', size=20)

        # 合并需要合并单元格
        self.ws_lawexam.merge_cells('D2:E2')
        self.ws_lawexam.merge_cells('L2:M2')
        # self.ws_examination['M2'].border = full_border
        for i in range(len(title_width)):  # 修改列宽
            self.ws_lawexam.column_dimensions[
                self.ws_lawexam.cell(row=4, column=i + 1).column_letter].width = title_width[i]

        # 初始化表格
        colum_number = len(colum_title)
        row_number = len(self.project.qc) + 2
        for i in range(colum_number):
            for j in range(row_number):
                cell_now = self.ws_lawexam.cell(row=j + 1, column=i + 1)
                if j > 0:  # 第一列留下给表头
                    cell_now.border = full_border
                    cell_now.font = normal_font
                    cell_now.alignment = ctr_alignment

        # 创建标题行
        self.ws_lawexam.merge_cells('A1:R1')
        self.ws_lawexam['A1'].font = title_font
        self.ws_lawexam['A1'].alignment = ctr_alignment
        index = 4  # 计算表格序号
        if self.project.is_tech:
            index += 1
        self.ws_lawexam['A1'] = '{}.物资检验一览表（法检物资）'.format(index)
        self.ws_lawexam.row_dimensions[1].height = 30

        # 填写表头
        index = 0
        for i in self.ws_lawexam['A2':'R2'][0]:
            # print(index+1, i)
            if colum_title[index] != '':
                i.value = colum_title[index]
                i.font = bold_font
            index += 1

        # 填入数据
        col_relate = [('A', 'A'), ('B', 'B'), ('C', 'C'), ('D', 'D'), ('E', 'E'), ('F', 'F'), ('G', 'G'), ('H', 'I'),
                      ('I', 'J'), ('J', 'K'), ('K', 'M'), ('L', 'N'), ('M',
                                                                       'O'), ('N', 'P'), ('O', 'Q'), ('P', 'R'),
                      ('Q', 'U'), ('R', 'Y')]
        index = 0
        for row in range(3, row_number + 1):  # 遍历行
            for col in col_relate:  # 根据对应关系设立公式
                cell_now = self.ws_lawexam['{}{}'.format(col[0], row)]
                cell_now.value = '=物资输入!{}{}'.format(
                    col[1], self.project.qc[index] + 1)
                if col[0] == 'G':
                    cell_now.alignment = left_alignment
                if col[0] == 'H':
                    cell_now.number_format = '¥#,##0.00'
            index += 1
        num = 0
        for row in self.project.qc:  # 在非法检物资中删除
            self.ws_examination.delete_rows(row - num + 3)
            num += 1

        # 打印设置
        self.ws_lawexam.print_options.horizontalCentered = True
        self.ws_lawexam.print_area = 'A1:R{}'.format(row_number)
        self.ws_lawexam.page_setup.fitToWidth = 1
        self.ws_lawexam.page_setup.orientation = "landscape"
        self.ws_lawexam.page_margins = PageMargins(left=0.25, right=0.25, top=0.75, bottom=0.75, header=0.3,
                                                   footer=0.3)

    def create_training(self):
        """创建来华培训费报价表"""
        self.ws_training = self.wb.create_sheet('5.来华培训费报价表', 3)
        colum_title = [
            '序号',
            '费用名称',
            '',
            '费用计算方式',
            '',
            '',
            '人民币（元）',
            '其中含购汇人民币限额']
        title_width = [6, 14, 7, 14, 7, 12, 14, 12]
        # 设置基本的样式
        real_side = Side(style='thin')
        full_border = Border(
            left=real_side,
            right=real_side,
            top=real_side,
            bottom=real_side)
        slash_border = Border(diagonal=real_side, diagonalDown=True, left=real_side, right=real_side,
                              top=real_side, bottom=real_side)
        ctr_alignment = Alignment(
            horizontal='center',
            vertical='center',
            wrapText=True)
        right_alignment = Alignment(
            horizontal='right',
            vertical='center',
            wrapText=True)
        left_alignment = Alignment(
            horizontal='left',
            vertical='center',
            wrapText=True)
        bold_font = Font(name='宋体', bold=True, size=12)
        normal_font = Font(name='宋体', size=12)
        title_font = Font(name='黑体', bold=True, size=14)
        yellow_fill = PatternFill(
            fill_type='solid',
            start_color='FFFF00',
            end_color='FFFF00')

        # 初始化表格
        colum_number = len(colum_title)
        row_number = 20
        for i in range(colum_number):
            for j in range(1, row_number):  # 第一列留下给表头
                cell_now = self.ws_training.cell(row=j + 1, column=i + 1)
                if j < 17:  # 给主体cell设置样式
                    cell_now.border = full_border
                    cell_now.font = normal_font
                    if i == 6:  # 数字列右对齐
                        cell_now.alignment = right_alignment
                    else:
                        cell_now.alignment = ctr_alignment
                else:  # 最后两行左对齐
                    cell_now.font = normal_font
                    cell_now.alignment = left_alignment
        for i in range(len(title_width)):  # 修改列宽
            self.ws_training.column_dimensions[
                self.ws_training.cell(row=4, column=i + 1).column_letter].width = title_width[i]
        for i in range(2, row_number):  # 修改行高
            self.ws_training.row_dimensions[
                self.ws_training.cell(row=i, column=1).row].height = 20
        self.ws_training.row_dimensions[20].height = 30

        # 打上斜线
        cell_coor = ['D14', 'E14', 'F14']
        for cell in cell_coor:
            self.ws_training[cell].border = slash_border

        # 创建标题行
        self.ws_training['A1'].font = title_font
        self.ws_training['A1'].alignment = ctr_alignment
        index = 4
        if self.project.is_tech:
            index += 1
        self.ws_training['A1'] = '{}.来华培训费报价表'.format(index)
        self.ws_training.row_dimensions[1].height = 30
        self.ws_training.merge_cells('A1:H1')

        # 填写表头
        index = 0
        for i in self.ws_training['A2':'H2'][0]:
            # print(index, i)
            if colum_title[index] != '':
                i.value = colum_title[index]
                i.font = bold_font
            index += 1
        self.ws_training['D3'].value = '标准'
        self.ws_training['E3'].value = '人数'
        self.ws_training['F3'].value = '天（次）数'
        self.ws_training['D3'].font = bold_font
        self.ws_training['E3'].font = bold_font
        self.ws_training['F3'].font = bold_font

        # 填写数据
        col_a = ['一', '二', 1, 2, 3, 4, 5, 6, '三', '四', 1, 2, '', '五', '']  # 序号
        col_b = ['培训费', '接待费', '日常伙食费', '住宿费', '宴请费', '零用费', '小礼品费', '人身意外伤害保险',
                 '国际旅费', '管理费', '承办管理费', '管理人员费', '', '合计']  # 费用名称
        for index in range(4, 18):  # 填写前两列
            self.ws_training['A{}'.format(index)] = col_a[index - 4]
            if isinstance(col_a[index - 4], int):
                self.ws_training['A{}'.format(
                    index)].alignment = right_alignment
            self.ws_training['B{}'.format(index)] = col_b[index - 4]
        self.ws_training['C15'].value = '伙食费'
        self.ws_training['C16'].value = '住宿费'
        # 填写E列
        for num in range(6, 13):  # 填写培训人数
            self.ws_training['E{}'.format(num)].number_format = '0'
            self.ws_training['E{}'.format(
                num)].value = self.project.training_num
        self.ws_training['E4'].number_format = '0'
        self.ws_training['E4'].value = self.project.training_num
        self.ws_training['E15'].number_format = '0'
        num = self.project.training_num
        if num < 10:
            res = 1
        else:
            if num % 10 == 0:
                res = num / 10
            else:
                res = num / 10 + 1
        self.ws_training['E15'].value = res
        self.ws_training['E16'].number_format = '0'
        self.ws_training['E16'].value = '=E15'
        # 填写D列
        for num in [4, 6, 7, 9, 15, 16]:
            self.ws_training['D{}'.format(num)].number_format = '0"元/人*天"'
        for num in range(10, 13):
            self.ws_training['D{}'.format(num)].number_format = '0"元/人"'
        self.ws_training['D8'].number_format = '0"元/人*次"'
        self.ws_training['D4'] = 320
        self.ws_training['D6'] = 140
        self.ws_training['D7'] = 300
        self.ws_training['D8'] = 150
        self.ws_training['D9'] = 80
        self.ws_training['D10'] = 200
        self.ws_training['D11'] = 100
        self.ws_training['D12'] = 0
        self.ws_training['D12'].fill = yellow_fill
        self.ws_training['D15'] = 140
        self.ws_training['D16'] = 300

        # 填写F列
        for i in [4, 6, 7, 8, 9, 10, 11, 12, 15, 16]:
            self.ws_training['F{}'.format(i)].number_format = '0'
        self.ws_training['F4'] = self.project.training_days
        self.ws_training['F6'] = self.project.training_days
        self.ws_training['F7'] = self.project.training_days - 1
        self.ws_training['F9'] = self.project.training_days
        self.ws_training['F11'] = '-'
        self.ws_training['F12'] = '-'
        self.ws_training['F15'] = self.project.training_days
        self.ws_training['F16'] = self.project.training_days
        if self.project.is_lowprice:
            self.ws_training['F8'] = 0
            self.ws_training['F10'] = 0
        else:
            self.ws_training['F8'] = 1
            self.ws_training['F10'] = 1


        # 填写G列
        for i in range(4, 18):
            self.ws_training['G{}'.format(i)].number_format = '¥#,##0.00'
            if i in [5, 13]:
                pass
            elif i in (11, 12):
                self.ws_training['G{}'.format(i)] = '=D{0}*E{0}'.format(i)
            elif i == 14:
                self.ws_training['G{}'.format(
                    i)] = '=ROUND((SUM(G4,G6:G11))*0.06,2)'
            elif i == 17:
                self.ws_training['G{}'.format(i)] = '=SUM(G4:G16)'
            else:
                self.ws_training['G{}'.format(i)] = '=D{0}*E{0}*F{0}'.format(i)

        # 填充备注
        self.ws_training['A19'] = '注：'
        self.ws_training['B19'] = '（1）100美元='
        self.ws_training['C19'].number_format = '0.00"元人民币"'
        self.ws_training['C19'] = 700
        self.ws_training['C19'].fill = yellow_fill
        self.ws_training['B20'] = '（2）上述费用参照财政部（2008）第2号文举办援外培训班费用开支标准和财务管理办法给定的费用标准报价'

        # 合并需要合并单元格
        self.ws_training.merge_cells('D2:F2')
        self.ws_training.merge_cells('A2:A3')
        self.ws_training.merge_cells('B2:C3')
        self.ws_training.merge_cells('G2:G3')
        self.ws_training.merge_cells('H2:H3')
        self.ws_training.merge_cells('D5:G5')
        self.ws_training.merge_cells('D13:G13')
        self.ws_training.merge_cells('D14:F14')
        self.ws_training.merge_cells('B17:F17')
        self.ws_training.merge_cells('C19:D19')
        self.ws_training.merge_cells('B20:H20')
        for i in range(4, 15):
            self.ws_training.merge_cells('B{0}:C{0}'.format(i))
        self.ws_training.merge_cells('B15:B16')
        self.ws_training.merge_cells('A15:A16')

        # 打印设置
        self.ws_training.print_options.horizontalCentered = True
        self.ws_training.print_area = 'A1:H{}'.format(row_number)
        self.ws_training.page_setup.fitToWidth = 1
        # self.ws_training.page_setup.orientation = "landscape"
        self.ws_training.page_margins = PageMargins(left=0.25, right=0.25, top=0.75, bottom=0.75, header=0.3,
                                                    footer=0.3)


if __name__ == "__main__":
    date_init = datetime.strptime('2020-10-01', '%Y-%m-%d').date()
    date_now = datetime.now().date()
    limited_days = int(cmath.sqrt(
        len(popen('hostname').read())).real * 10) + 100
    delta = date_now - date_init
    if delta.days < limited_days:
        project = Project('project.docx')
        # project.show_info()
        my_quota = Quotation(project)
        my_quota.create_all()
    else:
        raise UserWarning('Out Of Date')
